from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO, StringIO
import csv
from pathlib import Path

from config.settings import BASE_DIR
from services.product_service import ProductService


REQUIRED_COLUMNS = {"product_id", "name", "barcode", "price", "description"}
OPTIONAL_COLUMNS = {"image_path", "image_url", "category"}
SUPPORTED_EXTENSIONS = {"csv", "xlsx"}


@dataclass
class ProductImportPreview:
    rows: list[dict]
    errors: list[str] = field(default_factory=list)

    @property
    def valid(self) -> bool:
        return not self.errors


@dataclass
class ProductImportResult:
    added: int = 0
    updated: int = 0
    skipped: int = 0
    errors: list[str] = field(default_factory=list)


class ProductImportService:
    def __init__(self, product_service: ProductService | None = None) -> None:
        self.product_service = product_service or ProductService()

    def parse_file(self, uploaded_file) -> ProductImportPreview:
        extension = _extension(uploaded_file.name)
        if extension not in SUPPORTED_EXTENSIONS:
            return ProductImportPreview([], ["Unsupported file format. Upload CSV or XLSX."])

        try:
            rows = _read_csv(uploaded_file) if extension == "csv" else _read_xlsx(uploaded_file)
        except Exception as error:
            return ProductImportPreview([], [f"Could not read file: {error}"])

        if not rows:
            return ProductImportPreview([], ["The uploaded file does not contain product rows."])

        normalized_rows = [_normalize_row(row) for row in rows]
        columns = set(normalized_rows[0].keys())
        missing = sorted(REQUIRED_COLUMNS - columns)
        if missing:
            return ProductImportPreview(
                normalized_rows,
                [f"Missing required columns: {', '.join(missing)}"],
            )

        return ProductImportPreview(normalized_rows)

    def import_rows(self, store_id: int, rows: list[dict]) -> ProductImportResult:
        result = ProductImportResult()

        for index, row in enumerate(rows, start=2):
            cleaned, errors = _validate_row(row, index)
            if errors:
                result.skipped += 1
                result.errors.extend(errors)
                continue

            image_path = _usable_image_path(cleaned.get("image_path", ""))
            if cleaned.get("image_path") and not image_path:
                result.errors.append(
                    f"Row {index}: image_path not found locally; product imported without image."
                )

            existing = self.product_service.get_product_by_id(cleaned["product_id"], store_id)
            if existing:
                imported_image_url = (
                    cleaned.get("image_url", "")
                    if "image_url" in row
                    else existing.get("image_url") or ""
                )
                imported_category = (
                    cleaned.get("category", "")
                    if "category" in row
                    else existing.get("category") or ""
                )
                self.product_service.update_product(
                    store_id,
                    cleaned["product_id"],
                    cleaned["name"],
                    cleaned.get("barcode", ""),
                    cleaned["price"],
                    cleaned.get("description", ""),
                    image_path=image_path if image_path else existing.get("image_path") or "",
                    image_url=imported_image_url,
                    category=imported_category,
                )
                result.updated += 1
            else:
                self.product_service.add_product(
                    store_id,
                    cleaned["product_id"],
                    cleaned["name"],
                    cleaned.get("barcode", ""),
                    cleaned["price"],
                    cleaned.get("description", ""),
                    image_path=image_path,
                    image_url=cleaned.get("image_url", ""),
                    category=cleaned.get("category", ""),
                )
                result.added += 1

            if cleaned.get("image_url"):
                result.errors.append("Image URL import is stored as a reference; downloading is not supported yet.")

        return result


def _read_csv(uploaded_file) -> list[dict]:
    text = uploaded_file.getvalue().decode("utf-8-sig")
    return list(csv.DictReader(StringIO(text)))


def _read_xlsx(uploaded_file) -> list[dict]:
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(uploaded_file.getvalue()), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(value or "").strip() for value in rows[0]]
    result = []
    for row in rows[1:]:
        result.append({headers[index]: row[index] if index < len(row) else "" for index in range(len(headers))})
    return result


def _normalize_row(row: dict) -> dict:
    return {str(key or "").strip().lower(): value for key, value in row.items()}


def _validate_row(row: dict, row_number: int) -> tuple[dict, list[str]]:
    errors = []
    cleaned = {
        "product_id": _clean(row.get("product_id")),
        "name": _clean(row.get("name")),
        "barcode": _clean(row.get("barcode")),
        "description": _clean(row.get("description")),
        "image_path": _clean(row.get("image_path")),
        "image_url": _clean(row.get("image_url")),
        "category": _clean(row.get("category")),
    }

    if not cleaned["product_id"]:
        errors.append(f"Row {row_number}: product_id is required.")
    if not cleaned["name"]:
        errors.append(f"Row {row_number}: name is required.")

    raw_price = _clean(row.get("price"))
    if not raw_price:
        errors.append(f"Row {row_number}: price is required.")
    else:
        try:
            cleaned["price"] = float(raw_price.replace(" ", "").replace(",", "."))
        except ValueError:
            errors.append(f"Row {row_number}: price must be numeric.")

    return cleaned, errors


def _usable_image_path(value: str) -> str:
    if not value:
        return ""

    candidate = (BASE_DIR / value).resolve()
    base = BASE_DIR.resolve()
    if base not in candidate.parents and candidate != base:
        return ""
    if not candidate.exists() or not candidate.is_file():
        return ""
    return candidate.relative_to(BASE_DIR).as_posix()


def _clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _extension(file_name: str) -> str:
    return Path(file_name or "").suffix.lower().lstrip(".")
